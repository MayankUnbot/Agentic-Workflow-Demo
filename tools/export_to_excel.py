"""
Excel Exporter for Parsed Job Data
Exports parsed DailyRemote jobs to a formatted .xlsx file.

Usage:
    python tools/export_to_excel.py [--input .tmp/dailyremote/parsed_jobs.json] [--output .tmp/dailyremote/dailyremote_jobs.xlsx]
"""

import argparse
import json
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
DEFAULT_INPUT = os.path.join(BASE_DIR, ".tmp", "dailyremote", "parsed_jobs.json")
DEFAULT_OUTPUT = os.path.join(BASE_DIR, ".tmp", "dailyremote", "dailyremote_jobs.xlsx")

COLUMNS = [
    ("Job Title", "title", 45),
    ("Company", "company", 20),
    ("Employment Type", "employment_type", 16),
    ("Location", "location", 20),
    ("Salary", "salary", 25),
    ("Experience", "experience", 14),
    ("Category", "subcategory", 20),
    ("Description", "description", 60),
    ("Posted", "posted", 16),
    ("Job URL", "url", 50),
    ("Search Term", "search_term", 14),
]


def export_to_excel(jobs: list[dict], output_path: str):
    """Export job data to a formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Remote Jobs"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write headers
    for col_idx, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data rows
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for row_idx, job in enumerate(jobs, 2):
        for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
            value = job.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap_align

    # Freeze header row and add auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(jobs) + 1}"

    # Set row height for header
    ws.row_dimensions[1].height = 25

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Exported {len(jobs)} jobs to: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Export parsed jobs to Excel")
    parser.add_argument("--input", "-i", default=DEFAULT_INPUT, help="Path to parsed_jobs.json")
    parser.add_argument("--output", "-o", default=DEFAULT_OUTPUT, help="Output .xlsx path")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    with open(args.input, "r", encoding="utf-8") as f:
        jobs = json.load(f)

    print(f"Loaded {len(jobs)} jobs from: {args.input}")
    export_to_excel(jobs, args.output)


if __name__ == "__main__":
    main()
